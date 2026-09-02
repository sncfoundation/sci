#!/usr/bin/env python3
"""
sheetbuild — the reference SICF tool: pack an OCI container image INTO a spreadsheet
(the SICF native image store) and back out again, losslessly.

  sheetbuild import  <docker-save.tar> --name doom:shareware --store store.xlsx
  sheetbuild export  doom:shareware --store store.xlsx --out image.tar   # docker load < image.tar
  sheetbuild ls      --store store.xlsx
  sheetbuild info    doom:shareware --store store.xlsx

The store is a workbook laid out per the SICF v0.1 spec:
  Images tab: name, digest, config, layers, created, size
  Layers tab: digest, ordinal, media_type, data      (base64, sharded across cells)

Layers are content-addressed by sha256 and split into cell-sized base64 chunks so they
respect the spreadsheet cell-character limit (Excel 32,767 / Google Sheets 50,000; we
cap conservatively). On the way out, every layer is reassembled and its digest verified
before the image tar is rebuilt — a mismatch is a hard error.

This is the write/read path only; a runtime (kubelet) resolves `sicf:<name>` by calling
export/materialize and `docker load`-ing the result. Execution stays on the node — the
spreadsheet is the store and the control plane, never the executor.

Requires: openpyxl (stdlib otherwise). Produce the input with `docker save img -o x.tar`.
"""
import argparse, base64, hashlib, io, json, os, tarfile, time

MAX_CELL = int(os.environ.get("SICF_MAX_CELL", "30000"))   # safe for Excel (32767) and Sheets (50000)
IMAGES = "Images"
LAYERS = "Layers"
IMAGES_COLS = ["name", "digest", "config", "layers", "created", "size"]
LAYERS_COLS = ["digest", "ordinal", "media_type", "data"]
LAYER_MEDIA = "application/vnd.oci.image.layer.v1.tar"


def _sha256(b):
    return "sha256:" + hashlib.sha256(b).hexdigest()

def _shards(b64, n=None):
    n = n or MAX_CELL           # resolved at call time so the cap stays configurable
    return [b64[i:i + n] for i in range(0, len(b64), n)] or [""]

# ------------------------------------------------------------------ workbook

def _open(path):
    import openpyxl
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook(); wb.remove(wb.active)
    for tab, cols in ((IMAGES, IMAGES_COLS), (LAYERS, LAYERS_COLS)):
        if tab not in wb.sheetnames:
            wb.create_sheet(tab).append(cols)
    return wb

def _rows(ws):
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    head = [str(h) for h in data[0]]
    return [dict(zip(head, r)) for r in data[1:] if r and r[0] not in (None, "")]

def _has_layer(wb, digest):
    return any(r["digest"] == digest for r in _rows(wb[LAYERS]))

def _write_layer(wb, digest, blob):
    """Store a layer as ordered base64 chunks (skip if already present — content addressed)."""
    if _has_layer(wb, digest):
        return
    ws = wb[LAYERS]
    b64 = base64.b64encode(blob).decode()
    for ordinal, chunk in enumerate(_shards(b64)):
        ws.append([digest, ordinal, LAYER_MEDIA, chunk])

def _read_layer(wb, digest):
    """Reassemble a layer from its chunks (ordered by ordinal) and VERIFY its digest."""
    parts = sorted((r for r in _rows(wb[LAYERS]) if r["digest"] == digest),
                   key=lambda r: int(r["ordinal"]))
    if not parts:
        raise KeyError(f"layer {digest} not found in store")
    blob = base64.b64decode("".join(str(p["data"] or "") for p in parts))
    if _sha256(blob) != digest:
        raise ValueError(f"digest mismatch for {digest}: store is corrupt")
    return blob

# --------------------------------------------------------------- import/export

def import_image(tar_path, name, store_path):
    """Read a `docker save` tar and write it into the SICF store."""
    wb = _open(store_path)
    with tarfile.open(tar_path, "r") as tar:
        manifest = json.load(tar.extractfile("manifest.json"))
        entry = manifest[0]
        config_blob = tar.extractfile(entry["Config"]).read()
        if not name:
            tags = entry.get("RepoTags") or []
            name = tags[0] if tags else "imported:latest"
        layer_digests, total = [], 0
        for lpath in entry["Layers"]:
            blob = tar.extractfile(lpath).read()
            d = _sha256(blob)
            _write_layer(wb, d, blob)
            layer_digests.append(d); total += len(blob)
    digest = _sha256(config_blob)
    # upsert Images row
    ws = wb[IMAGES]
    row_vals = [name, digest, base64.b64encode(config_blob).decode(),
                ",".join(layer_digests), time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()), total]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == name:
            for i, v in enumerate(row_vals):
                row[i].value = v
            break
    else:
        ws.append(row_vals)
    wb.save(store_path)
    return {"name": name, "digest": digest, "layers": len(layer_digests), "bytes": total}

def _image_row(wb, name):
    for r in _rows(wb[IMAGES]):
        if r["name"] == name:
            return r
    raise KeyError(f"image {name} not found in store")

def export_image(name, store_path, out_tar):
    """Rebuild a `docker load`-compatible tar from the store (verifies every layer)."""
    wb = _open(store_path)
    img = _image_row(wb, name)
    config_blob = base64.b64decode(img["config"])
    config_name = hashlib.sha256(config_blob).hexdigest() + ".json"
    layer_digests = [d for d in str(img["layers"]).split(",") if d]

    with tarfile.open(out_tar, "w") as tar:
        def _add(arcname, data):
            ti = tarfile.TarInfo(arcname); ti.size = len(data)
            tar.addfile(ti, io.BytesIO(data))
        _add(config_name, config_blob)
        layer_paths = []
        for d in layer_digests:
            blob = _read_layer(wb, d)                       # raises on digest mismatch
            path = d.split(":", 1)[1] + "/layer.tar"
            _add(path, blob); layer_paths.append(path)
        manifest = [{"Config": config_name, "RepoTags": [name], "Layers": layer_paths}]
        _add("manifest.json", json.dumps(manifest).encode())
    return {"name": name, "layers": len(layer_paths), "out": out_tar}

# ------------------------------------------------------------------------ CLI

def main():
    ap = argparse.ArgumentParser(description="SICF reference tool: OCI image <-> spreadsheet")
    sub = ap.add_subparsers(dest="cmd")
    p = sub.add_parser("import", help="import a `docker save` tar into the store")
    p.add_argument("tar"); p.add_argument("--name", default=""); p.add_argument("--store", required=True)
    p = sub.add_parser("export", help="rebuild a docker-loadable tar from the store")
    p.add_argument("name"); p.add_argument("--store", required=True); p.add_argument("--out", required=True)
    p = sub.add_parser("ls", help="list images in the store"); p.add_argument("--store", required=True)
    p = sub.add_parser("info", help="show one image"); p.add_argument("name"); p.add_argument("--store", required=True)
    a = ap.parse_args()

    if a.cmd == "import":
        print(json.dumps(import_image(a.tar, a.name, a.store)))
    elif a.cmd == "export":
        print(json.dumps(export_image(a.name, a.store, a.out)))
    elif a.cmd == "ls":
        for r in _rows(_open(a.store)[IMAGES]):
            print(f'{r["name"]:24} {r["digest"][:19]}  {r["size"]} bytes  {len(str(r["layers"]).split(","))} layers')
    elif a.cmd == "info":
        print(json.dumps(_image_row(_open(a.store), a.name), indent=2, default=str))
    else:
        ap.print_help()

if __name__ == "__main__":
    main()
